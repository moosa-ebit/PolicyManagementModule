import openpyxl, io
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Policy, PolicyAcknowledgment

def adjustColumnWidth(worksheet):
    column_widths = []
    for row in worksheet.rows:
        for i, cell in enumerate(row):
            cell_value = cell.value or ''
            if len(column_widths) > i:
                if len(cell_value) > column_widths[i]:
                    column_widths[i] = len(cell_value)
            else:
                column_widths += [len(cell_value)]
    
    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        worksheet.column_dimensions[get_column_letter(i)].width = column_width

def generatePolicyAcknowledgementReport(policy: Policy):
    font_size = 11
    cell_font = Font(bold=True, size=font_size)
    blue_fill = PatternFill(start_color='00B4D8', end_color='00B4D8', fill_type='solid')
    
    acknowledgements = PolicyAcknowledgment.objects.filter(policy = policy)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Policy", policy.title])
    sheet.append([])
    sheet['A1'].fill = blue_fill
    sheet['A1'].font = cell_font
    sheet.append(["First Name", "Last Name", "Time of Acknowledgment"])
    for cell in ['A3', 'B3', 'C3']:
        sheet[cell].fill = blue_fill
        sheet[cell].font = cell_font
    for ack in acknowledgements:
        sheet.append([ack.user.first_name, ack.user.last_name, ack.acknowledged_at.strftime(("%d/%m/%Y %H:%M %p"))])

    adjustColumnWidth(sheet)
    buffer = io.BytesIO()
    workbook.save(buffer)

    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=policy_acknowledgement_report (' + policy.title + ').xlsx'
    return response